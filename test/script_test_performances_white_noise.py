################################################################################
#HEADER

#This script allows you to test the performances of the BWE, PBWE and SSBWE
#packages facing different levels of white-noise, to check for regressions.
#It will generate a report containing the test results.

#The test consists in measuring the performances of the BWE, PBWE and SSBWE on a
#radar signal containing 2 echoes from 2 targets, separated by distances below
#or close to the radar's resolution.

#The following test parameters are user-defined:
#   -The list of distances between targets to be tested (m).
#   -The list of SNR to be tested (dB).
#   -The number of noise cases

#The following tests will be performed:
#   -The percentage of times 2 echoes are detected (with a user-defined
#    threshold).
#   -The error on the distance between targets, estimated from the surface echo.
#   -The error on the each echo's amplitude.

#The synthetic radar signal example (inspired by the WISDOM GPR of the ExoMars
#rover mission, Ciarletti et al. (2017)):
#   -A SFCW (Stepped Frequency Continuous Wave) radar working between 0.5 and 3
#    GHz measures a 1001 frequencies spectrum when sounding.
#   -Only the In-phase component (real part of the spectrum) is measured, the
#    Quadrature component (imaginary part of the spectrum) is reconstructed by
#    Hilbert transform.
#   -Two targets in free-space are seperated by 5 cm, slightly below the radar's
#    free-space resolution. These targets generate echoes of equal amplitudes in
#    the radar's signal, or complex sine-waves in the measured spectrum.
#   -The measured spectrum is corrupted by a white-noise of standard deviation
#    10X smaller than the complex sine-waves' amplitudes.

#The parameters used for the BWE, PBWE and SSBWE are the default ones.

#References: Oudart et al. (2021)

################################################################################

#Libraries importation:---------------------------------------------------------

import numpy as np
import pandas as pd
import os
from math import pi
from scipy.signal import hilbert,find_peaks
from openpyxl.styles import Font

import PyBWE
import PyPBWE
import PySSBWE

#User-defined test parameters:--------------------------------------------------

#List of distances between targets to be tested (m):
list_dist_targets = [0.04,0.06,0.08,0.1,0.12]

#List of SNR (dB):
list_snr_levels = [6,14,20,40,60]

#Number of noise case to be tested for each distance:
nb_noise_case = 100

#Other parameters:--------------------------------------------------------------

#Generate a vector of 1001 frequencies between 0.5 and 3 GHz:
freq_vect = np.linspace(0.5e9,3e9,1001)

#Amplitude of the 2 echoes corresponding to the 2 targets:
amp_target1 = 1
amp_target2 = 1

#Distance (m) between the 1st radar target in free-space (returning one echo)
#and the radar:
dist_target1 = 1

#Peak detection threshold on the amplitude of echoes:
detection_level = 0.1

#Retrieve the test path:--------------------------------------------------------

#Test directory path:
test_dir_path = os.path.dirname(__file__)

#Test report path:
test_report_path = os.path.join(test_dir_path,'Report_test_performances_white_noise.xlsx')

#Initialize the test results dataframes:----------------------------------------

#For the BWE:
dataframe_bwe_percentage_echoes_detection = pd.DataFrame(0,index=list_snr_levels,columns=list_dist_targets)
dataframe_bwe_distance_mean_error = pd.DataFrame(0,index=list_snr_levels,columns=list_dist_targets)
dataframe_bwe_distance_std_error = pd.DataFrame(0,index=list_snr_levels,columns=list_dist_targets)
dataframe_bwe_amplitude_1_mean_error = pd.DataFrame(0,index=list_snr_levels,columns=list_dist_targets)
dataframe_bwe_amplitude_1_std_error = pd.DataFrame(0,index=list_snr_levels,columns=list_dist_targets)
dataframe_bwe_amplitude_2_mean_error = pd.DataFrame(0,index=list_snr_levels,columns=list_dist_targets)
dataframe_bwe_amplitude_2_std_error = pd.DataFrame(0,index=list_snr_levels,columns=list_dist_targets)

#For the PBWE - polar 00:
dataframe_pbwe_polar00_percentage_echoes_detection = pd.DataFrame(0,index=list_snr_levels,columns=list_dist_targets)
dataframe_pbwe_polar00_distance_mean_error = pd.DataFrame(0,index=list_snr_levels,columns=list_dist_targets)
dataframe_pbwe_polar00_distance_std_error = pd.DataFrame(0,index=list_snr_levels,columns=list_dist_targets)
dataframe_pbwe_polar00_amplitude_1_mean_error = pd.DataFrame(0,index=list_snr_levels,columns=list_dist_targets)
dataframe_pbwe_polar00_amplitude_1_std_error = pd.DataFrame(0,index=list_snr_levels,columns=list_dist_targets)
dataframe_pbwe_polar00_amplitude_2_mean_error = pd.DataFrame(0,index=list_snr_levels,columns=list_dist_targets)
dataframe_pbwe_polar00_amplitude_2_std_error = pd.DataFrame(0,index=list_snr_levels,columns=list_dist_targets)

#For the PBWE - polar 11:
dataframe_pbwe_polar11_percentage_echoes_detection = pd.DataFrame(0,index=list_snr_levels,columns=list_dist_targets)
dataframe_pbwe_polar11_distance_mean_error = pd.DataFrame(0,index=list_snr_levels,columns=list_dist_targets)
dataframe_pbwe_polar11_distance_std_error = pd.DataFrame(0,index=list_snr_levels,columns=list_dist_targets)
dataframe_pbwe_polar11_amplitude_1_mean_error = pd.DataFrame(0,index=list_snr_levels,columns=list_dist_targets)
dataframe_pbwe_polar11_amplitude_1_std_error = pd.DataFrame(0,index=list_snr_levels,columns=list_dist_targets)
dataframe_pbwe_polar11_amplitude_2_mean_error = pd.DataFrame(0,index=list_snr_levels,columns=list_dist_targets)
dataframe_pbwe_polar11_amplitude_2_std_error = pd.DataFrame(0,index=list_snr_levels,columns=list_dist_targets)

#For the SSBWE - method 1:
dataframe_ssbwe_method1_percentage_echoes_detection = pd.DataFrame(0,index=list_snr_levels,columns=list_dist_targets)
dataframe_ssbwe_method1_distance_mean_error = pd.DataFrame(0,index=list_snr_levels,columns=list_dist_targets)
dataframe_ssbwe_method1_distance_std_error = pd.DataFrame(0,index=list_snr_levels,columns=list_dist_targets)
dataframe_ssbwe_method1_amplitude_1_mean_error = pd.DataFrame(0,index=list_snr_levels,columns=list_dist_targets)
dataframe_ssbwe_method1_amplitude_1_std_error = pd.DataFrame(0,index=list_snr_levels,columns=list_dist_targets)
dataframe_ssbwe_method1_amplitude_2_mean_error = pd.DataFrame(0,index=list_snr_levels,columns=list_dist_targets)
dataframe_ssbwe_method1_amplitude_2_std_error = pd.DataFrame(0,index=list_snr_levels,columns=list_dist_targets)

#For the SSBWE - method 2:
dataframe_ssbwe_method2_percentage_echoes_detection = pd.DataFrame(0,index=list_snr_levels,columns=list_dist_targets)
dataframe_ssbwe_method2_distance_mean_error = pd.DataFrame(0,index=list_snr_levels,columns=list_dist_targets)
dataframe_ssbwe_method2_distance_std_error = pd.DataFrame(0,index=list_snr_levels,columns=list_dist_targets)
dataframe_ssbwe_method2_amplitude_1_mean_error = pd.DataFrame(0,index=list_snr_levels,columns=list_dist_targets)
dataframe_ssbwe_method2_amplitude_1_std_error = pd.DataFrame(0,index=list_snr_levels,columns=list_dist_targets)
dataframe_ssbwe_method2_amplitude_2_mean_error = pd.DataFrame(0,index=list_snr_levels,columns=list_dist_targets)
dataframe_ssbwe_method2_amplitude_2_std_error = pd.DataFrame(0,index=list_snr_levels,columns=list_dist_targets)


#Perform the test:--------------------------------------------------------------

#Initialize the iterations counter:
nb_iterations = len(list_dist_targets)*len(list_snr_levels)*nb_noise_case
count_iterations = 0

#Iterate on the distance between targets:
for dist in list_dist_targets:

    #Distance (m) between the 2nd radar target in free-space (returning one
    #echo) and the radar:
    dist_target2 = 1 + dist

    #Generate a sum of two complex sine-waves corresponding to the targets' echoes:
    spec_vect_00 = (amp_target1*np.exp(-1j*4*pi*dist_target1*freq_vect/3e8))+(amp_target2*np.exp(-1j*4*pi*dist_target2*freq_vect/3e8))
    spec_vect_11 = (amp_target1*np.exp(-1j*4*pi*dist_target1*freq_vect/3e8))-(amp_target2*np.exp(-1j*4*pi*dist_target2*freq_vect/3e8))

    #Only keep the real part of the spectrum (In-phase component):
    spec_vect_00 = np.real(spec_vect_00)
    spec_vect_11 = np.real(spec_vect_11)

    #Calculate the power of the signal (dB):
    power_sig_00 = 10*np.log10(np.mean(spec_vect_00**2))
    power_sig_11 = 10*np.log10(np.mean(spec_vect_11**2))

    #Iterate on the SNR values:
    for snr in list_snr_levels:

        #Initialize the 2 echoes detection counter:
        nb_echoes_detection_bwe = 0
        nb_echoes_detection_pbwe00 = 0
        nb_echoes_detection_pbwe11 = 0
        nb_echoes_detection_ssbwe1 = 0
        nb_echoes_detection_ssbwe2 = 0

        #Initialize the lists to stock all noise cases results:
        list_distance_bwe = []
        list_amplitude_target1_bwe = []
        list_amplitude_target2_bwe = []
        list_distance_pbwe00 = []
        list_amplitude_target1_pbwe00 = []
        list_amplitude_target2_pbwe00 = []
        list_distance_pbwe11 = []
        list_amplitude_target1_pbwe11 = []
        list_amplitude_target2_pbwe11 = []
        list_distance_ssbwe1 = []
        list_amplitude_target1_ssbwe1 = []
        list_amplitude_target2_ssbwe1 = []
        list_distance_ssbwe2 = []
        list_amplitude_target1_ssbwe2 = []
        list_amplitude_target2_ssbwe2 = []

        #Calculate the power of the noise (dB) to obtain the right SNR:
        power_noise_00 = power_sig_00-snr
        power_noise_11 = power_sig_11-snr

        #Iterate on the noise case:
        for idx_case in range(nb_noise_case):

            #Display the test progress:
            count_iterations += 1
            print('Test in progress ... '+str(round(100*count_iterations/nb_iterations,3))+'%')

            #Add a white noise to this spectrum signal to obtain the right SNR:
            wn_vect_00 = np.random.normal(0,np.sqrt(10**(power_noise_00/10)),spec_vect_00.shape)
            wn_vect_11 = np.random.normal(0,np.sqrt(10**(power_noise_11/10)),spec_vect_11.shape)
            spec_vect_wn_00 = spec_vect_00 + wn_vect_00
            spec_vect_wn_11 = spec_vect_11 + wn_vect_11

            #Reconstruct a complex signal with the Hilbert transform:
            spec_vect_wn_00 = np.conjugate(hilbert(spec_vect_wn_00))[::2]
            spec_vect_wn_11 = np.conjugate(hilbert(spec_vect_wn_11))[::2]

            #Assemble the 2 spectrum channels into a single matrix:
            spec_mat_wn = np.vstack((spec_vect_wn_00,spec_vect_wn_11))

            #Application of the BWE:
            output_bwe, time_bwe_vect = PyBWE.BWE(spec_vect_wn_00,df=5e6,extra_factor=3,model_order=0.33,zp_factor=10,side_cut=True)

            #Application of the PBWE:
            output_pbwe, time_pbwe_vect = PyPBWE.PBWE(spec_mat_wn,df=5e6,extra_factor=3,model_order=0.33,zp_factor=10,side_cut=True)

            #Application of the SSBWE:
            output_ssbwe_1, output_ssbwe_2, time_ssbwe_vect = PySSBWE.SSBWE(spec_vect_wn_00,df=5e6,extra_factor=3,zp_factor=10,side_cut=True)

            #Testing the BWE:---------------------------------------------------

            #Detection of the 2 echoes:
            peaks_bwe = find_peaks(abs(output_bwe),height=detection_level)[0]

            #Check that 2 echoes are detected:
            if len(peaks_bwe)==2:

                #Increment by 1 the number of echoes detection:
                nb_echoes_detection_bwe += 1

                #Stock the distance between echoes:
                list_distance_bwe += [0.5*abs(time_bwe_vect[peaks_bwe[0]]-time_bwe_vect[peaks_bwe[1]])*3e8]

                #Stock the amplitude of the 1st echo:
                list_amplitude_target1_bwe += [abs(output_bwe[peaks_bwe[0]])]

                #Stock the amplitude of the 2nd echo:
                list_amplitude_target2_bwe += [abs(output_bwe[peaks_bwe[1]])]

            #Testing the PBWE - polar 00:---------------------------------------

            #Detection of the 2 echoes:
            peaks_pbwe_00 = find_peaks(abs(output_pbwe[0,:]),height=detection_level)[0]

            #Check that 2 echoes are detected:
            if len(peaks_pbwe_00)==2:

                #Increment by 1 the number of echoes detection:
                nb_echoes_detection_pbwe00 += 1

                #Stock the distance between echoes:
                list_distance_pbwe00 += [0.5*abs(time_pbwe_vect[peaks_pbwe_00[0]]-time_pbwe_vect[peaks_pbwe_00[1]])*3e8]

                #Stock the amplitude of the 1st echo:
                list_amplitude_target1_pbwe00 += [abs(output_pbwe[0,peaks_pbwe_00[0]])]

                #Stock the amplitude of the 2nd echo:
                list_amplitude_target2_pbwe00 += [abs(output_pbwe[0,peaks_pbwe_00[1]])]

            #Testing the PBWE - polar 11:---------------------------------------

            #Detection of the 2 echoes:
            peaks_pbwe_11 = find_peaks(abs(output_pbwe[1,:]),height=detection_level)[0]

            #Check that 2 echoes are detected:
            if len(peaks_pbwe_11)==2:

                #Increment by 1 the number of echoes detection:
                nb_echoes_detection_pbwe11 += 1

                #Stock the distance between echoes:
                list_distance_pbwe11 += [0.5*abs(time_pbwe_vect[peaks_pbwe_11[0]]-time_pbwe_vect[peaks_pbwe_11[1]])*3e8]

                #Stock the amplitude of the 1st echo:
                list_amplitude_target1_pbwe11 += [abs(output_pbwe[1,peaks_pbwe_11[0]])]

                #Stock the amplitude of the 2nd echo:
                list_amplitude_target2_pbwe11 += [abs(output_pbwe[1,peaks_pbwe_11[1]])]

            #Testing the SSBWE - method 1:--------------------------------------

            #Detection of the 2 echoes:
            peaks_ssbwe_1 = find_peaks(abs(output_ssbwe_1),height=detection_level)[0]

            #Check that 2 echoes are detected:
            if len(peaks_ssbwe_1)==2:

                #Increment by 1 the number of echoes detection:
                nb_echoes_detection_ssbwe1 += 1

                #Stock the distance between echoes:
                list_distance_ssbwe1 += [0.5*abs(time_ssbwe_vect[peaks_ssbwe_1[0]]-time_ssbwe_vect[peaks_ssbwe_1[1]])*3e8]

                #Stock the amplitude of the 1st echo:
                list_amplitude_target1_ssbwe1 += [abs(output_ssbwe_1[peaks_ssbwe_1[0]])]

                #Stock the amplitude of the 2nd echo:
                list_amplitude_target2_ssbwe1 += [abs(output_ssbwe_1[peaks_ssbwe_1[1]])]

            #Testing the SSBWE - method 2:--------------------------------------

            #Detection of the 2 echoes:
            peaks_ssbwe_2 = find_peaks(abs(output_ssbwe_2),height=detection_level)[0]

            #Check that 2 echoes are detected:
            if len(peaks_ssbwe_2)==2:

                #Increment by 1 the number of echoes detection:
                nb_echoes_detection_ssbwe2 += 1

                #Stock the distance between echoes:
                list_distance_ssbwe2 += [0.5*abs(time_ssbwe_vect[peaks_ssbwe_2[0]]-time_ssbwe_vect[peaks_ssbwe_2[1]])*3e8]

                #Stock the amplitude of the 1st echo:
                list_amplitude_target1_ssbwe2 += [abs(output_ssbwe_2[peaks_ssbwe_2[0]])]

                #Stock the amplitude of the 2nd echo:
                list_amplitude_target2_ssbwe2 += [abs(output_ssbwe_2[peaks_ssbwe_2[1]])]


        #Tests calculations:----------------------------------------------------

        #Calculate the percentage of echoes detection:
        pc_echoes_detection_bwe = 100*nb_echoes_detection_bwe/nb_noise_case
        pc_echoes_detection_pbwe00 = 100*nb_echoes_detection_pbwe00/nb_noise_case
        pc_echoes_detection_pbwe11 = 100*nb_echoes_detection_pbwe11/nb_noise_case
        pc_echoes_detection_ssbwe1 = 100*nb_echoes_detection_ssbwe1/nb_noise_case
        pc_echoes_detection_ssbwe2 = 100*nb_echoes_detection_ssbwe2/nb_noise_case

        #Calculate the error on the distance between echoes:
        error_distance_bwe = np.array(list_distance_bwe)-dist
        error_distance_pbwe00 = np.array(list_distance_pbwe00)-dist
        error_distance_pbwe11 = np.array(list_distance_pbwe11)-dist
        error_distance_ssbwe1 = np.array(list_distance_ssbwe1)-dist
        error_distance_ssbwe2 = np.array(list_distance_ssbwe2)-dist

        #Calculate the error on the amplitude of the 1st echo:
        error_amplitude_target1_bwe = np.array(list_amplitude_target1_bwe)-amp_target1
        error_amplitude_target1_pbwe00 = np.array(list_amplitude_target1_pbwe00)-amp_target1
        error_amplitude_target1_pbwe11 = np.array(list_amplitude_target1_pbwe11)-amp_target1
        error_amplitude_target1_ssbwe1 = np.array(list_amplitude_target1_ssbwe1)-amp_target1
        error_amplitude_target1_ssbwe2 = np.array(list_amplitude_target1_ssbwe2)-amp_target1

        #Calculate the error on the amplitude of the 2nd echo:
        error_amplitude_target2_bwe = np.array(list_amplitude_target2_bwe)-amp_target2
        error_amplitude_target2_pbwe00 = np.array(list_amplitude_target2_pbwe00)-amp_target2
        error_amplitude_target2_pbwe11 = np.array(list_amplitude_target2_pbwe11)-amp_target2
        error_amplitude_target2_ssbwe1 = np.array(list_amplitude_target2_ssbwe1)-amp_target2
        error_amplitude_target2_ssbwe2 = np.array(list_amplitude_target2_ssbwe2)-amp_target2

        #Add the tests results to the dataframes:-------------------------------

        #Add the BWE test results to the corresponding dataframes:
        dataframe_bwe_percentage_echoes_detection.loc[snr,dist] = pc_echoes_detection_bwe
        dataframe_bwe_distance_mean_error.loc[snr,dist] = np.mean(error_distance_bwe)
        dataframe_bwe_distance_std_error.loc[snr,dist] = np.std(error_distance_bwe)
        dataframe_bwe_amplitude_1_mean_error.loc[snr,dist] = np.mean(error_amplitude_target1_bwe)
        dataframe_bwe_amplitude_1_std_error.loc[snr,dist] = np.std(error_amplitude_target1_bwe)
        dataframe_bwe_amplitude_2_mean_error.loc[snr,dist] = np.mean(error_amplitude_target2_bwe)
        dataframe_bwe_amplitude_2_std_error.loc[snr,dist] = np.std(error_amplitude_target2_bwe)

        #Add the PBWE - polar 00 test results to the corresponding dataframes:
        dataframe_pbwe_polar00_percentage_echoes_detection.loc[snr,dist] = pc_echoes_detection_pbwe00
        dataframe_pbwe_polar00_distance_mean_error.loc[snr,dist] = np.mean(error_distance_pbwe00)
        dataframe_pbwe_polar00_distance_std_error.loc[snr,dist] = np.std(error_distance_pbwe00)
        dataframe_pbwe_polar00_amplitude_1_mean_error.loc[snr,dist] = np.mean(error_amplitude_target1_pbwe00)
        dataframe_pbwe_polar00_amplitude_1_std_error.loc[snr,dist] = np.std(error_amplitude_target1_pbwe00)
        dataframe_pbwe_polar00_amplitude_2_mean_error.loc[snr,dist] = np.mean(error_amplitude_target2_pbwe00)
        dataframe_pbwe_polar00_amplitude_2_std_error.loc[snr,dist] = np.std(error_amplitude_target2_pbwe00)

        #Add the PBWE - polar 11 test results to the corresponding dataframes:
        dataframe_pbwe_polar11_percentage_echoes_detection.loc[snr,dist] = pc_echoes_detection_pbwe11
        dataframe_pbwe_polar11_distance_mean_error.loc[snr,dist] = np.mean(error_distance_pbwe11)
        dataframe_pbwe_polar11_distance_std_error.loc[snr,dist] = np.std(error_distance_pbwe11)
        dataframe_pbwe_polar11_amplitude_1_mean_error.loc[snr,dist] = np.mean(error_amplitude_target1_pbwe11)
        dataframe_pbwe_polar11_amplitude_1_std_error.loc[snr,dist] = np.std(error_amplitude_target1_pbwe11)
        dataframe_pbwe_polar11_amplitude_2_mean_error.loc[snr,dist] = np.mean(error_amplitude_target2_pbwe11)
        dataframe_pbwe_polar11_amplitude_2_std_error.loc[snr,dist] = np.std(error_amplitude_target2_pbwe11)

        #The SSBWE - method 1 test results to the corresponding dataframes:
        dataframe_ssbwe_method1_percentage_echoes_detection.loc[snr,dist] = pc_echoes_detection_ssbwe1
        dataframe_ssbwe_method1_distance_mean_error.loc[snr,dist] = np.mean(error_distance_ssbwe1)
        dataframe_ssbwe_method1_distance_std_error.loc[snr,dist] = np.std(error_distance_ssbwe1)
        dataframe_ssbwe_method1_amplitude_1_mean_error.loc[snr,dist] = np.mean(error_amplitude_target1_ssbwe1)
        dataframe_ssbwe_method1_amplitude_1_std_error.loc[snr,dist] = np.std(error_amplitude_target1_ssbwe1)
        dataframe_ssbwe_method1_amplitude_2_mean_error.loc[snr,dist] = np.mean(error_amplitude_target2_ssbwe1)
        dataframe_ssbwe_method1_amplitude_2_std_error.loc[snr,dist] = np.std(error_amplitude_target2_ssbwe1)

        #The SSBWE - method 2 test results to the corresponding dataframes:
        dataframe_ssbwe_method2_percentage_echoes_detection.loc[snr,dist] = pc_echoes_detection_ssbwe2
        dataframe_ssbwe_method2_distance_mean_error.loc[snr,dist] = np.mean(error_distance_ssbwe2)
        dataframe_ssbwe_method2_distance_std_error.loc[snr,dist] = np.std(error_distance_ssbwe2)
        dataframe_ssbwe_method2_amplitude_1_mean_error.loc[snr,dist] = np.mean(error_amplitude_target1_ssbwe2)
        dataframe_ssbwe_method2_amplitude_1_std_error.loc[snr,dist] = np.std(error_amplitude_target1_ssbwe2)
        dataframe_ssbwe_method2_amplitude_2_mean_error.loc[snr,dist] = np.mean(error_amplitude_target2_ssbwe2)
        dataframe_ssbwe_method2_amplitude_2_std_error.loc[snr,dist] = np.std(error_amplitude_target2_ssbwe2)

#Export the Excel report:-------------------------------------------------------

with pd.ExcelWriter(test_report_path) as writer:

    #Excel sheets corresponding to the BWE tests:-------------------------------

    dataframe_bwe_percentage_echoes_detection.to_excel(writer, sheet_name='BWE_echoes_detection',index_label='SNR (dB)',startrow = 1)
    worksheet = writer.sheets['BWE_echoes_detection']
    worksheet['B1'] = 'Distance between targets (m)'
    worksheet['B1'].font = Font(bold=True)
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(list_dist_targets)+1)

    dataframe_bwe_distance_mean_error.to_excel(writer, sheet_name='BWE_mean_distance_error',index_label='SNR (dB)',startrow = 1)
    worksheet = writer.sheets['BWE_mean_distance_error']
    worksheet['B1'] = 'Distance between targets (m)'
    worksheet['B1'].font = Font(bold=True)
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(list_dist_targets)+1)

    dataframe_bwe_distance_std_error.to_excel(writer, sheet_name='BWE_STD_distance_error',index_label='SNR (dB)',startrow = 1)
    worksheet = writer.sheets['BWE_STD_distance_error']
    worksheet['B1'] = 'Distance between targets (m)'
    worksheet['B1'].font = Font(bold=True)
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(list_dist_targets)+1)

    dataframe_bwe_amplitude_1_mean_error.to_excel(writer, sheet_name='BWE_mean_amplitude_error_1',index_label='SNR (dB)',startrow = 1)
    worksheet = writer.sheets['BWE_mean_amplitude_error_1']
    worksheet['B1'] = 'Distance between targets (m)'
    worksheet['B1'].font = Font(bold=True)
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(list_dist_targets)+1)

    dataframe_bwe_amplitude_1_std_error.to_excel(writer, sheet_name='BWE_STD_amplitude_error_1',index_label='SNR (dB)',startrow = 1)
    worksheet = writer.sheets['BWE_STD_amplitude_error_1']
    worksheet['B1'] = 'Distance between targets (m)'
    worksheet['B1'].font = Font(bold=True)
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(list_dist_targets)+1)

    dataframe_bwe_amplitude_2_mean_error.to_excel(writer, sheet_name='BWE_mean_amplitude_error_2',index_label='SNR (dB)',startrow = 1)
    worksheet = writer.sheets['BWE_mean_amplitude_error_2']
    worksheet['B1'] = 'Distance between targets (m)'
    worksheet['B1'].font = Font(bold=True)
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(list_dist_targets)+1)

    dataframe_bwe_amplitude_2_std_error.to_excel(writer, sheet_name='BWE_STD_amplitude_error_2',index_label='SNR (dB)',startrow = 1)
    worksheet = writer.sheets['BWE_STD_amplitude_error_2']
    worksheet['B1'] = 'Distance between targets (m)'
    worksheet['B1'].font = Font(bold=True)
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(list_dist_targets)+1)

    #Excel sheets corresponding to the PBWE - polar 00 tests:-------------------

    dataframe_pbwe_polar00_percentage_echoes_detection.to_excel(writer, sheet_name='PBWE_00_echoes_detection',index_label='SNR (dB)',startrow = 1)
    worksheet = writer.sheets['PBWE_00_echoes_detection']
    worksheet['B1'] = 'Distance between targets (m)'
    worksheet['B1'].font = Font(bold=True)
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(list_dist_targets)+1)

    dataframe_pbwe_polar00_distance_mean_error.to_excel(writer, sheet_name='PBWE_00_mean_distance_error',index_label='SNR (dB)',startrow = 1)
    worksheet = writer.sheets['PBWE_00_mean_distance_error']
    worksheet['B1'] = 'Distance between targets (m)'
    worksheet['B1'].font = Font(bold=True)
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(list_dist_targets)+1)

    dataframe_pbwe_polar00_distance_std_error.to_excel(writer, sheet_name='PBWE_00_STD_distance_error',index_label='SNR (dB)',startrow = 1)
    worksheet = writer.sheets['PBWE_00_STD_distance_error']
    worksheet['B1'] = 'Distance between targets (m)'
    worksheet['B1'].font = Font(bold=True)
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(list_dist_targets)+1)

    dataframe_pbwe_polar00_amplitude_1_mean_error.to_excel(writer, sheet_name='PBWE_00_mean_amplitude_error_1',index_label='SNR (dB)',startrow = 1)
    worksheet = writer.sheets['PBWE_00_mean_amplitude_error_1']
    worksheet['B1'] = 'Distance between targets (m)'
    worksheet['B1'].font = Font(bold=True)
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(list_dist_targets)+1)

    dataframe_pbwe_polar00_amplitude_1_std_error.to_excel(writer, sheet_name='PBWE_00_STD_amplitude_error_1',index_label='SNR (dB)',startrow = 1)
    worksheet = writer.sheets['PBWE_00_STD_amplitude_error_1']
    worksheet['B1'] = 'Distance between targets (m)'
    worksheet['B1'].font = Font(bold=True)
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(list_dist_targets)+1)

    dataframe_pbwe_polar00_amplitude_2_mean_error.to_excel(writer, sheet_name='PBWE_00_mean_amplitude_error_2',index_label='SNR (dB)',startrow = 1)
    worksheet = writer.sheets['PBWE_00_mean_amplitude_error_2']
    worksheet['B1'] = 'Distance between targets (m)'
    worksheet['B1'].font = Font(bold=True)
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(list_dist_targets)+1)

    dataframe_pbwe_polar00_amplitude_2_std_error.to_excel(writer, sheet_name='PBWE_00_STD_amplitude_error_2',index_label='SNR (dB)',startrow = 1)
    worksheet = writer.sheets['PBWE_00_STD_amplitude_error_2']
    worksheet['B1'] = 'Distance between targets (m)'
    worksheet['B1'].font = Font(bold=True)
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(list_dist_targets)+1)

    #Excel sheets corresponding to the PBWE - polar 11 tests:-------------------

    dataframe_pbwe_polar11_percentage_echoes_detection.to_excel(writer, sheet_name='PBWE_11_echoes_detection',index_label='SNR (dB)',startrow = 1)
    worksheet = writer.sheets['PBWE_11_echoes_detection']
    worksheet['B1'] = 'Distance between targets (m)'
    worksheet['B1'].font = Font(bold=True)
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(list_dist_targets)+1)

    dataframe_pbwe_polar11_distance_mean_error.to_excel(writer, sheet_name='PBWE_11_mean_distance_error',index_label='SNR (dB)',startrow = 1)
    worksheet = writer.sheets['PBWE_11_mean_distance_error']
    worksheet['B1'] = 'Distance between targets (m)'
    worksheet['B1'].font = Font(bold=True)
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(list_dist_targets)+1)

    dataframe_pbwe_polar11_distance_std_error.to_excel(writer,sheet_name='PBWE_11_STD_distance_error',index_label='SNR (dB)',startrow = 1)
    worksheet = writer.sheets['PBWE_11_STD_distance_error']
    worksheet['B1'] = 'Distance between targets (m)'
    worksheet['B1'].font = Font(bold=True)
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(list_dist_targets)+1)

    dataframe_pbwe_polar11_amplitude_1_mean_error.to_excel(writer, sheet_name='PBWE_11_mean_amplitude_error_1',index_label='SNR (dB)',startrow = 1)
    worksheet = writer.sheets['PBWE_11_mean_amplitude_error_1']
    worksheet['B1'] = 'Distance between targets (m)'
    worksheet['B1'].font = Font(bold=True)
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(list_dist_targets)+1)

    dataframe_pbwe_polar11_amplitude_1_std_error.to_excel(writer, sheet_name='PBWE_11_STD_amplitude_error_1',index_label='SNR (dB)',startrow = 1)
    worksheet = writer.sheets['PBWE_11_STD_amplitude_error_1']
    worksheet['B1'] = 'Distance between targets (m)'
    worksheet['B1'].font = Font(bold=True)
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(list_dist_targets)+1)

    dataframe_pbwe_polar11_amplitude_2_mean_error.to_excel(writer, sheet_name='PBWE_11_mean_amplitude_error_2',index_label='SNR (dB)',startrow = 1)
    worksheet = writer.sheets['PBWE_11_mean_amplitude_error_2']
    worksheet['B1'] = 'Distance between targets (m)'
    worksheet['B1'].font = Font(bold=True)
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(list_dist_targets)+1)

    dataframe_pbwe_polar11_amplitude_2_std_error.to_excel(writer, sheet_name='PBWE_11_STD_amplitude_error_2',index_label='SNR (dB)',startrow = 1)
    worksheet = writer.sheets['PBWE_11_STD_amplitude_error_2']
    worksheet['B1'] = 'Distance between targets (m)'
    worksheet['B1'].font = Font(bold=True)
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(list_dist_targets)+1)

    #Excel sheets corresponding to the SSBWE - method 1 tests:------------------

    dataframe_ssbwe_method1_percentage_echoes_detection.to_excel(writer, sheet_name='SSBWE_v1_echoes_detection',index_label='SNR (dB)',startrow = 1)
    worksheet = writer.sheets['SSBWE_v1_echoes_detection']
    worksheet['B1'] = 'Distance between targets (m)'
    worksheet['B1'].font = Font(bold=True)
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(list_dist_targets)+1)

    dataframe_ssbwe_method1_distance_mean_error.to_excel(writer, sheet_name='SSBWE_v1_mean_distance_error',index_label='SNR (dB)',startrow = 1)
    worksheet = writer.sheets['SSBWE_v1_mean_distance_error']
    worksheet['B1'] = 'Distance between targets (m)'
    worksheet['B1'].font = Font(bold=True)
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(list_dist_targets)+1)

    dataframe_ssbwe_method1_distance_std_error.to_excel(writer, sheet_name='SSBWE_v1_STD_distance_error',index_label='SNR (dB)',startrow = 1)
    worksheet = writer.sheets['SSBWE_v1_STD_distance_error']
    worksheet['B1'] = 'Distance between targets (m)'
    worksheet['B1'].font = Font(bold=True)
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(list_dist_targets)+1)

    dataframe_ssbwe_method1_amplitude_1_mean_error.to_excel(writer, sheet_name='SSBWE_v1_mean_amplitude_error_1',index_label='SNR (dB)',startrow = 1)
    worksheet = writer.sheets['SSBWE_v1_mean_amplitude_error_1']
    worksheet['B1'] = 'Distance between targets (m)'
    worksheet['B1'].font = Font(bold=True)
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(list_dist_targets)+1)

    dataframe_ssbwe_method1_amplitude_1_std_error.to_excel(writer, sheet_name='SSBWE_v1_STD_amplitude_error_1',index_label='SNR (dB)',startrow = 1)
    worksheet = writer.sheets['SSBWE_v1_STD_amplitude_error_1']
    worksheet['B1'] = 'Distance between targets (m)'
    worksheet['B1'].font = Font(bold=True)
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(list_dist_targets)+1)

    dataframe_ssbwe_method1_amplitude_2_mean_error.to_excel(writer, sheet_name='SSBWE_v1_mean_amplitude_error_2',index_label='SNR (dB)',startrow = 1)
    worksheet = writer.sheets['SSBWE_v1_mean_amplitude_error_2']
    worksheet['B1'] = 'Distance between targets (m)'
    worksheet['B1'].font = Font(bold=True)
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(list_dist_targets)+1)

    dataframe_ssbwe_method1_amplitude_2_std_error.to_excel(writer, sheet_name='SSBWE_v1_STD_amplitude_error_2',index_label='SNR (dB)',startrow = 1)
    worksheet = writer.sheets['SSBWE_v1_STD_amplitude_error_2']
    worksheet['B1'] = 'Distance between targets (m)'
    worksheet['B1'].font = Font(bold=True)
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(list_dist_targets)+1)

    #Excel sheets corresponding to the SSBWE - method 2 tests:------------------

    dataframe_ssbwe_method2_percentage_echoes_detection.to_excel(writer, sheet_name='SSBWE_v2_echoes_detection',index_label='SNR (dB)',startrow = 1)
    worksheet = writer.sheets['SSBWE_v2_echoes_detection']
    worksheet['B1'] = 'Distance between targets (m)'
    worksheet['B1'].font = Font(bold=True)
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(list_dist_targets)+1)

    dataframe_ssbwe_method2_distance_mean_error.to_excel(writer, sheet_name='SSBWE_v2_mean_distance_error',index_label='SNR (dB)',startrow = 1)
    worksheet = writer.sheets['SSBWE_v2_mean_distance_error']
    worksheet['B1'] = 'Distance between targets (m)'
    worksheet['B1'].font = Font(bold=True)
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(list_dist_targets)+1)

    dataframe_ssbwe_method2_distance_std_error.to_excel(writer, sheet_name='SSBWE_v2_STD_distance_error',index_label='SNR (dB)',startrow = 1)
    worksheet = writer.sheets['SSBWE_v2_STD_distance_error']
    worksheet['B1'] = 'Distance between targets (m)'
    worksheet['B1'].font = Font(bold=True)
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(list_dist_targets)+1)

    dataframe_ssbwe_method2_amplitude_1_mean_error.to_excel(writer, sheet_name='SSBWE_v2_mean_amplitude_error_1',index_label='SNR (dB)',startrow = 1)
    worksheet = writer.sheets['SSBWE_v2_mean_amplitude_error_1']
    worksheet['B1'] = 'Distance between targets (m)'
    worksheet['B1'].font = Font(bold=True)
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(list_dist_targets)+1)

    dataframe_ssbwe_method2_amplitude_1_std_error.to_excel(writer, sheet_name='SSBWE_v2_STD_amplitude_error_1',index_label='SNR (dB)',startrow = 1)
    worksheet = writer.sheets['SSBWE_v2_STD_amplitude_error_1']
    worksheet['B1'] = 'Distance between targets (m)'
    worksheet['B1'].font = Font(bold=True)
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(list_dist_targets)+1)

    dataframe_ssbwe_method2_amplitude_2_mean_error.to_excel(writer, sheet_name='SSBWE_v2_mean_amplitude_error_2',index_label='SNR (dB)',startrow = 1)
    worksheet = writer.sheets['SSBWE_v2_mean_amplitude_error_2']
    worksheet['B1'] = 'Distance between targets (m)'
    worksheet['B1'].font = Font(bold=True)
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(list_dist_targets)+1)

    dataframe_ssbwe_method2_amplitude_2_std_error.to_excel(writer, sheet_name='SSBWE_v2_STD_amplitude_error_2',index_label='SNR (dB)',startrow = 1)
    worksheet = writer.sheets['SSBWE_v2_STD_amplitude_error_2']
    worksheet['B1'] = 'Distance between targets (m)'
    worksheet['B1'].font = Font(bold=True)
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(list_dist_targets)+1)
